"""
Excel file handler for financial data ingestion.
Handles XLSX and XLS files with multiple sheets.
"""
import pandas as pd
import io
from typing import Dict, List, Optional, Tuple
from datetime import datetime
from openpyxl import load_workbook
import re


class ExcelHandler:
    """Handles Excel file parsing and data extraction."""
    
    # Sheet name patterns for different financial statements
    SHEET_PATTERNS = {
        'income_statement': ['income', 'p&l', 'profit', 'loss', 'revenue', 'pl_statement'],
        'balance_sheet': ['balance', 'bs', 'assets', 'liabilities', 'position'],
        'cash_flow': ['cash', 'cashflow', 'cash_flow', 'cf'],
        'trial_balance': ['trial', 'tb', 'trial_balance'],
        'transactions': ['transactions', 'ledger', 'journal', 'entries'],
        'receivables': ['receivable', 'ar', 'debtors', 'customers'],
        'payables': ['payable', 'ap', 'creditors', 'vendors'],
        'inventory': ['inventory', 'stock', 'items'],
    }
    
    def __init__(self):
        self.errors = []
        self.warnings = []
    
    def parse_file(
        self, 
        file_content: bytes, 
        filename: str
    ) -> Tuple[Dict[str, pd.DataFrame], Dict]:
        """
        Parse an Excel file and return structured data.
        
        Args:
            file_content: Raw file bytes
            filename: Original filename
        
        Returns:
            Tuple of (dict of sheet_name: DataFrame, metadata dict)
        """
        self.errors = []
        self.warnings = []
        metadata = {
            'filename': filename,
            'format': 'xlsx' if filename.endswith('.xlsx') else 'xls',
            'parsed_at': datetime.utcnow().isoformat(),
            'sheets': {},
            'sheet_types': {}
        }
        
        sheets = {}
        
        try:
            # Load workbook to get sheet names
            workbook = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
            sheet_names = workbook.sheetnames
            workbook.close()
            
            # Read each sheet
            excel_file = pd.ExcelFile(io.BytesIO(file_content))
            
            for sheet_name in sheet_names:
                try:
                    df = pd.read_excel(excel_file, sheet_name=sheet_name)
                    
                    # Skip empty sheets
                    if df.empty:
                        self.warnings.append(f"Sheet '{sheet_name}' is empty")
                        continue
                    
                    # Clean column names
                    df.columns = [self._clean_column_name(str(col)) for col in df.columns]
                    
                    # Detect sheet type
                    sheet_type = self._detect_sheet_type(sheet_name, df)
                    
                    # Process data based on type
                    df = self._process_sheet(df, sheet_type)
                    
                    sheets[sheet_name] = df
                    metadata['sheets'][sheet_name] = {
                        'type': sheet_type,
                        'rows': len(df),
                        'columns': df.columns.tolist()
                    }
                    metadata['sheet_types'][sheet_name] = sheet_type
                    
                except Exception as e:
                    self.errors.append(f"Error reading sheet '{sheet_name}': {str(e)}")
            
            metadata['total_sheets'] = len(sheets)
            metadata['errors'] = self.errors
            metadata['warnings'] = self.warnings
            
            return sheets, metadata
            
        except Exception as e:
            self.errors.append(f"Failed to parse Excel file: {str(e)}")
            metadata['errors'] = self.errors
            return {}, metadata
    
    def _clean_column_name(self, name: str) -> str:
        """Clean and normalize column name."""
        # Handle unnamed columns
        if 'unnamed' in name.lower():
            return name
        # Remove extra whitespace, convert to lowercase
        name = re.sub(r'\s+', '_', name.strip().lower())
        # Remove special characters except underscore
        name = re.sub(r'[^a-z0-9_]', '', name)
        return name if name else 'column'
    
    def _detect_sheet_type(self, sheet_name: str, df: pd.DataFrame) -> str:
        """Detect the type of financial statement from sheet name and content."""
        sheet_lower = sheet_name.lower().replace(' ', '_')
        
        # Check sheet name patterns
        for stmt_type, patterns in self.SHEET_PATTERNS.items():
            for pattern in patterns:
                if pattern in sheet_lower:
                    return stmt_type
        
        # Check column content for clues
        columns_lower = ' '.join(df.columns.astype(str)).lower()
        
        if any(term in columns_lower for term in ['revenue', 'sales', 'expense', 'profit']):
            return 'income_statement'
        elif any(term in columns_lower for term in ['asset', 'liability', 'equity']):
            return 'balance_sheet'
        elif any(term in columns_lower for term in ['cash_inflow', 'cash_outflow', 'operating']):
            return 'cash_flow'
        elif any(term in columns_lower for term in ['debit', 'credit', 'balance']):
            return 'trial_balance'
        elif any(term in columns_lower for term in ['invoice', 'bill']):
            return 'transactions'
        
        return 'unknown'
    
    def _process_sheet(self, df: pd.DataFrame, sheet_type: str) -> pd.DataFrame:
        """Process sheet based on its type."""
        # Remove completely empty rows
        df = df.dropna(how='all')
        
        # Reset index
        df = df.reset_index(drop=True)
        
        # Convert numeric columns
        df = self._convert_numerics(df)
        
        # Convert date columns
        df = self._convert_dates(df)
        
        return df
    
    def _convert_numerics(self, df: pd.DataFrame) -> pd.DataFrame:
        """Convert numeric columns, handling currency formatting."""
        for col in df.columns:
            if df[col].dtype == 'object':
                try:
                    # Check if column looks numeric
                    sample = df[col].dropna().head(10)
                    if len(sample) > 0:
                        # Remove currency symbols and commas
                        cleaned = sample.astype(str).str.replace(r'[₹$,\s()]', '', regex=True)
                        # Check if most values are numeric
                        numeric_count = pd.to_numeric(cleaned, errors='coerce').notna().sum()
                        if numeric_count / len(sample) > 0.7:
                            df[col] = df[col].astype(str).str.replace(r'[₹$,\s]', '', regex=True)
                            # Handle parentheses for negative numbers
                            df[col] = df[col].str.replace(r'\((.+)\)', r'-\1', regex=True)
                            df[col] = pd.to_numeric(df[col], errors='coerce')
                except Exception:
                    pass
        return df
    
    def _convert_dates(self, df: pd.DataFrame) -> pd.DataFrame:
        """Convert date columns to datetime."""
        date_patterns = ['date', 'period', 'month', 'year', 'quarter']
        
        for col in df.columns:
            if any(pattern in col.lower() for pattern in date_patterns):
                try:
                    df[col] = pd.to_datetime(df[col], errors='coerce')
                except Exception:
                    pass
        return df
    
    def extract_financial_data(
        self, 
        sheets: Dict[str, pd.DataFrame], 
        metadata: Dict
    ) -> Dict:
        """
        Extract structured financial data from all sheets.
        
        Args:
            sheets: Dictionary of sheet DataFrames
            metadata: Metadata from parsing
        
        Returns:
            Consolidated financial data dictionary
        """
        financial_data = {
            'income_statement': None,
            'balance_sheet': None,
            'cash_flow': None,
            'trial_balance': None,
            'receivables': None,
            'payables': None,
            'inventory': None,
            'other': []
        }
        
        for sheet_name, df in sheets.items():
            sheet_type = metadata.get('sheet_types', {}).get(sheet_name, 'unknown')
            
            if sheet_type in financial_data and financial_data[sheet_type] is None:
                financial_data[sheet_type] = self._extract_sheet_data(df, sheet_type)
            elif sheet_type == 'unknown':
                financial_data['other'].append({
                    'sheet_name': sheet_name,
                    'data': self._extract_sheet_data(df, 'generic')
                })
        
        return financial_data
    
    def _extract_sheet_data(self, df: pd.DataFrame, sheet_type: str) -> Dict:
        """Extract data from a specific sheet."""
        # Convert to records, handling NaN and datetime
        records = []
        for _, row in df.iterrows():
            record = {}
            for col, val in row.items():
                if pd.isna(val):
                    record[col] = None
                elif isinstance(val, pd.Timestamp):
                    record[col] = val.isoformat()
                else:
                    record[col] = val
            records.append(record)
        
        # Calculate summary statistics
        numeric_cols = df.select_dtypes(include=['number']).columns
        summary = {}
        for col in numeric_cols:
            summary[col] = {
                'sum': float(df[col].sum()) if not df[col].isna().all() else None,
                'mean': float(df[col].mean()) if not df[col].isna().all() else None,
                'min': float(df[col].min()) if not df[col].isna().all() else None,
                'max': float(df[col].max()) if not df[col].isna().all() else None
            }
        
        return {
            'type': sheet_type,
            'row_count': len(df),
            'columns': df.columns.tolist(),
            'records': records,
            'summary': summary
        }


# Create global instance
excel_handler = ExcelHandler()
